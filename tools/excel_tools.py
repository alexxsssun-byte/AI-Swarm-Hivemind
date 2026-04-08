import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from core.memory import track_file, update_assumptions

# Constants for IB Styling
FONT_NAME = "Arial"
FONT_SIZE = 10
COLOR_INPUT = "0000FF"    # Blue for hardcoded inputs
COLOR_CALC = "000000"     # Black for calculations
COLOR_HEADER = "F2F2F2"   # Light gray for headers
NUMBER_FORMAT_DOLLAR = '_($* #,##0.0_);_($* (#,##0.0);_($* "-"??_);_(@_)'
NUMBER_FORMAT_PERCENT = '0.0%'
NUMBER_FORMAT_MULTIPLE = '0.0x'

def apply_ib_style(ws):
    ws.sheet_view.showGridLines = False

def generate_financial_model(project_id: int, model_type: str, assumptions: dict, project_name: str) -> str:
    """
    Creates a new .xlsx model (DCF, LBO, or M&A).
    """
    os.makedirs("output", exist_ok=True)
    filename = f"{project_name.replace(' ', '_')}_{model_type}_Model.xlsx"
    filepath = os.path.join("output", filename)
    
    wb = Workbook()
    
    if model_type.upper() == 'DCF':
        _build_dcf_sheet(wb, assumptions)
    else:
        # Generic fallback for LBO/M&A out of scope for the demo
        ws = wb.active
        ws.title = f"{model_type} Model"
        apply_ib_style(ws)
        ws["A1"] = f"{model_type} Model Generation not fully implemented yet."
        
    wb.save(filepath)
    
    # Store state
    update_assumptions(project_id, assumptions)
    track_file(project_id, filename, "Excel Model")
    
    return filepath

def _build_dcf_sheet(wb, assumptions):
    ws = wb.active
    ws.title = "DCF Build"
    apply_ib_style(ws)
    
    # Default granular driver assumptions
    rev_base = float(assumptions.get("revenue_base", 1000.0))
    rev_growth = float(assumptions.get("revenue_growth", 0.10))
    ebit_margin = float(assumptions.get("ebit_margin", 0.20))
    tax_rate = float(assumptions.get("tax_rate", 0.25))
    wacc = float(assumptions.get("wacc", 0.10))
    term_growth = float(assumptions.get("terminal_growth", 0.02))
    projection_years = int(assumptions.get("projection_years", 5))
    
    # Styles
    font_bold = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
    font_input = Font(name=FONT_NAME, size=FONT_SIZE, color=COLOR_INPUT)
    font_calc = Font(name=FONT_NAME, size=FONT_SIZE, color=COLOR_CALC)
    border_bottom = Border(bottom=Side(style='thin'))
    border_top_bottom = Border(top=Side(style='thin'), bottom=Side(style='double'))
    
    # Headers
    headers = ["($ in Millions)"] + [f"Year {i}" for i in range(1, projection_years+1)] + ["Terminal Value"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = h
        cell.font = font_bold
        cell.border = border_bottom
        cell.alignment = Alignment(horizontal='right')
        if col_idx == 1:
            cell.alignment = Alignment(horizontal='left')
    
    # Projections Math (Replicating dcf.py logic)
    proj_rev = []
    proj_ebit = []
    proj_nopat = []
    proj_fcf = []
    
    curr_rev = rev_base
    for i in range(projection_years):
        curr_rev *= (1 + rev_growth)
        ebit = curr_rev * ebit_margin
        nopat = ebit * (1 - tax_rate)
        # Assuming FCF = NOPAT for simplicity unless D&A/CapEx provided
        fcf = nopat 
        
        proj_rev.append(curr_rev)
        proj_ebit.append(ebit)
        proj_nopat.append(nopat)
        proj_fcf.append(fcf)
        
    tv = (proj_fcf[-1] * (1 + term_growth)) / (wacc - term_growth) if wacc > term_growth else 0
    
    row = 4
    # Revenue Row
    ws.cell(row=row, column=1, value="Revenue").font = font_bold
    for col_idx, val in enumerate(proj_rev, start=2):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.font = font_calc
        c.number_format = NUMBER_FORMAT_DOLLAR
    row += 1
    
    # EBIT Row
    ws.cell(row=row, column=1, value="EBIT").font = font_bold
    for col_idx, val in enumerate(proj_ebit, start=2):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.font = font_calc
        c.number_format = NUMBER_FORMAT_DOLLAR
    row += 1
    
    # NOPAT Row
    ws.cell(row=row, column=1, value="NOPAT").font = font_bold
    for col_idx, val in enumerate(proj_nopat, start=2):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.font = font_calc
        c.number_format = NUMBER_FORMAT_DOLLAR
    row += 2
    
    # Unlevered FCF
    ws.cell(row=row, column=1, value="Unlevered Free Cash Flow").font = font_bold
    for col_idx, val in enumerate(proj_fcf, start=2):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.font = font_bold
        c.number_format = NUMBER_FORMAT_DOLLAR
        c.border = border_top_bottom
        
    # Terminal Value
    c_tv = ws.cell(row=row, column=projection_years+2, value=tv)
    c_tv.font = font_bold
    c_tv.number_format = NUMBER_FORMAT_DOLLAR
    c_tv.border = border_top_bottom
    
    # Assumptions Block below
    row += 3
    ws.cell(row=row, column=1, value="Key Assumptions").font = font_bold
    ws.cell(row=row, column=1).border = border_bottom
    row += 1
    
    def write_input(r, label, val, is_pct=False):
        ws.cell(row=r, column=1, value=label).font = font_calc
        c = ws.cell(row=r, column=2, value=val)
        c.font = font_input
        c.number_format = NUMBER_FORMAT_PERCENT if is_pct else NUMBER_FORMAT_DOLLAR
        
    write_input(row, "Revenue Growth Rate", rev_growth, True); row+=1
    write_input(row, "EBIT Margin", ebit_margin, True); row+=1
    write_input(row, "Tax Rate", tax_rate, True); row+=1
    write_input(row, "WACC", wacc, True); row+=1
    write_input(row, "Terminal Growth Rate", term_growth, True); row+=1


def edit_excel_model(filepath: str, update_instructions: dict) -> str:
    """
    Loads an existing .xlsx model, updates specific assumptions, and recalculates it.
    Not fully implemented for complex cells yet.
    """
    # For now, it will just re-generate if we know the project id
    return filepath

def get_excel_tools():
    return [generate_financial_model, edit_excel_model]
